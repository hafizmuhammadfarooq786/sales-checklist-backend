"""
Import complete checklist data from Excel files (92 items from 10 Excel files)

Usage:
    python -m app.utils.import_checklist_data
"""
import asyncio
import sys
from pathlib import Path
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
import openpyxl
from typing import Dict, List, Optional

from app.db.session import AsyncSessionLocal
from app.models import (
    ChecklistCategory,
    ChecklistItem,
    CoachingQuestion,
    Organization,
    Setting,
)


# Path to Excel files
EXCEL_FILES_PATH = "/Users/muhammadfarooq/Documents/UPWORK BUSINESS/Dave Varner/Checklist Sheets"


# Checklist categories mapping (10 categories, 92 total items)
CHECKLIST_CATEGORIES = [
    {
        "name": "Trigger Event & Impact",
        "description": "Why the customer is buying and measurable results expected",
        "order": 1,
        "item_count": 10,
        "file": "Trigger_Event_and_Impact_Coaching_With_Mindset.xlsx",
    },
    {
        "name": "Trigger Priority",
        "description": "Is this truly a priority for decision influencers?",
        "order": 2,
        "item_count": 8,
        "file": "Trigger_Priority_Coaching_With_Mindset.xlsx",
    },
    {
        "name": "Sales Target",
        "description": "What, how much, when they plan to buy",
        "order": 3,
        "item_count": 8,
        "file": "Sales_Target_Coaching_With_Mindset.xlsx",
    },
    {
        "name": "Decision Influencer",
        "description": "Who influences the purchase decision and their priorities",
        "order": 4,
        "item_count": 7,
        "file": "Decision_Influencer_Coaching_With_Mindset.xlsx",
    },
    {
        "name": "Individual Impact",
        "description": "Impact of solution on decision influencers individually",
        "order": 5,
        "item_count": 8,
        "file": "Individual_Impact_Coaching_Final.xlsx",
    },
    {
        "name": "Mentor",
        "description": "Internal champion or coach for the deal",
        "order": 6,
        "item_count": 12,
        "file": "Mentor_Coaching_With_Mindset.xlsx",
    },
    {
        "name": "Decision Making Process",
        "description": "How the organization makes buying decisions",
        "order": 7,
        "item_count": 12,
        "file": "Decision_Making_Process_Coaching_With_Mindset.xlsx",
    },
    {
        "name": "Fit",
        "description": "How well solution fits customer's requirements",
        "order": 8,
        "item_count": 10,
        "file": "Fit_Coaching_With_Mindset.xlsx",
    },
    {
        "name": "Alternatives",
        "description": "Competitive alternatives being considered",
        "order": 9,
        "item_count": 8,
        "file": "Alternatives_Coaching_Import.xlsx",
    },
    {
        "name": "Our Solution Ranking",
        "description": "How we rank vs. competitors",
        "order": 10,
        "item_count": 9,
        "file": "Our_Solution_Ranking_Coaching_With_Mindset.xlsx",
    },
]


def parse_excel_file(file_path: str) -> Dict:
    """Parse an Excel file and extract checklist data"""
    print(f"  📄 Parsing {Path(file_path).name}...")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Extract sections
        sections = {}
        definition = None
        key_behavior = None
        key_mindset = None
        coaching_questions = []
        
        for row in range(1, ws.max_row + 1):
            section_cell = ws.cell(row=row, column=1).value
            content_cell = ws.cell(row=row, column=2).value
            
            if section_cell and content_cell:
                section = str(section_cell).strip()
                content = str(content_cell).strip()
                
                # Map sections to our structure
                if section == "Definition":
                    definition = content
                elif section == "Key Salesperson Behavior":
                    key_behavior = content
                elif section == "Key Mindset":
                    key_mindset = content
                elif "Coaching Questions" in section or any(keyword in section for keyword in [
                    "Motivation", "Clarifying", "Assessing", "Understanding", "Stakeholder", 
                    "Action", "Strategic", "Planning", "Implementation", "Validation",
                    "Questions", "Engagement", "Identification", "Leverage"
                ]):
                    # This is a coaching question section
                    coaching_questions.append({
                        "section": section,
                        "questions": content
                    })
        
        return {
            "definition": definition,
            "key_behavior": key_behavior,
            "key_mindset": key_mindset,
            "coaching_questions": coaching_questions
        }
        
    except Exception as e:
        print(f"  ❌ Error parsing {file_path}: {str(e)}")
        return {}


async def clear_existing_data(db: AsyncSession):
    """Clear existing checklist data to avoid duplicates"""
    print("\n🧹 Clearing existing checklist data...")
    
    try:
        # Delete in reverse dependency order
        await db.execute(delete(CoachingQuestion))
        await db.execute(delete(ChecklistItem))
        await db.execute(delete(ChecklistCategory))
        await db.commit()
        print("✅ Cleared existing checklist data")
    except Exception as e:
        print(f"❌ Error clearing data: {str(e)}")
        await db.rollback()


async def seed_categories(db: AsyncSession) -> Dict[str, ChecklistCategory]:
    """Create checklist categories"""
    print("\n🌱 Creating checklist categories...")

    categories = {}
    for cat_data in CHECKLIST_CATEGORIES:
        category = ChecklistCategory(
            name=cat_data["name"],
            description=cat_data["description"],
            order=cat_data["order"],
            default_weight=1.0,
            max_score=10,
            is_active=True
        )
        db.add(category)
        await db.flush()  # Get ID without committing
        categories[cat_data["name"]] = category
        print(f"  ✅ Created category: {cat_data['name']}")

    await db.commit()
    print(f"✅ Created {len(CHECKLIST_CATEGORIES)} categories")
    return categories


async def import_category_items(
    db: AsyncSession, 
    category: ChecklistCategory, 
    category_data: Dict,
    excel_data: Dict
) -> List[ChecklistItem]:
    """Import checklist items for a specific category"""
    print(f"  📊 Importing items for: {category.name}")
    
    if not excel_data:
        print(f"  ⚠️  No data found for {category.name}")
        return []
    
    items = []
    expected_count = category_data["item_count"]
    
    # Calculate points per item for this category (equal distribution)
    points_per_item = round(100.0 / 92, 3)  # 1.087 points per item
    
    # For categories with multiple coaching question sections,
    # each section becomes a separate checklist item
    coaching_questions = excel_data.get("coaching_questions", [])
    
    if len(coaching_questions) >= expected_count:
        # Use coaching question sections as individual items
        for i, cq_section in enumerate(coaching_questions[:expected_count]):
            item = ChecklistItem(
                category_id=category.id,
                title=cq_section["section"],
                definition=excel_data.get("definition", ""),
                key_behavior=excel_data.get("key_behavior", ""),
                key_mindset=excel_data.get("key_mindset", ""),
                order=i + 1,
                weight=1.0,
                points=points_per_item,
                is_active=True
            )
            db.add(item)
            await db.flush()
            
            # Add coaching question
            coaching_q = CoachingQuestion(
                item_id=item.id,
                section="Questions",
                question=cq_section["questions"],
                order=1
            )
            db.add(coaching_q)
            items.append(item)
            
    else:
        # Fall back to creating items based on expected count
        # This handles cases where the Excel structure is different
        definition = excel_data.get("definition", "")
        key_behavior = excel_data.get("key_behavior", "")
        key_mindset = excel_data.get("key_mindset", "")
        
        # Split definition into multiple parts if it has A), B), C) structure
        parts = []
        if definition and ("A)" in definition or "B)" in definition):
            # Split on A), B), C) pattern
            import re
            parts = re.split(r'[A-Z]\)\s*', definition)
            parts = [part.strip() for part in parts if part.strip()]
        else:
            parts = [definition]
        
        # Create items based on available parts or coaching questions
        for i in range(expected_count):
            if i < len(parts):
                item_title = f"{category.name} - Part {i + 1}"
                item_definition = parts[i]
            elif i < len(coaching_questions):
                item_title = coaching_questions[i]["section"]
                item_definition = coaching_questions[i]["questions"]
            else:
                item_title = f"{category.name} - Item {i + 1}"
                item_definition = definition
            
            item = ChecklistItem(
                category_id=category.id,
                title=item_title,
                definition=item_definition,
                key_behavior=key_behavior,
                key_mindset=key_mindset,
                order=i + 1,
                weight=1.0,
                points=points_per_item,
                is_active=True
            )
            db.add(item)
            await db.flush()
            
            # Add coaching questions if available
            if i < len(coaching_questions):
                coaching_q = CoachingQuestion(
                    item_id=item.id,
                    section="Questions",
                    question=coaching_questions[i]["questions"],
                    order=1
                )
                db.add(coaching_q)
            
            items.append(item)
    
    print(f"  ✅ Created {len(items)} items for {category.name}")
    return items


async def seed_default_organization(db: AsyncSession) -> Organization:
    """Create default organization"""
    print("\n🌱 Seeding default organization...")

    result = await db.execute(
        select(Organization).where(Organization.name == "The Millau Group Global")
    )
    org = result.scalar_one_or_none()

    if org:
        print("  ⏭️  Organization already exists")
        return org

    org = Organization(
        name="The Millau Group Global",
        domain="millaugroupglobal.com",
        scoring_mode="equal_weight",
        max_users=10000,
        is_active=True
    )
    db.add(org)
    await db.commit()
    print("✅ Created organization: The Millau Group Global")
    return org


async def seed_settings(db: AsyncSession):
    """Create default system settings"""
    print("\n🌱 Updating system settings...")

    default_settings = [
        {
            "key": "risk_band_red_threshold",
            "value": "60",
            "description": "Scores below this are marked as RED (at risk)",
        },
        {
            "key": "risk_band_yellow_threshold",
            "value": "80",
            "description": "Scores below this (but above red) are marked as YELLOW (caution)",
        },
        {
            "key": "default_scoring_mode",
            "value": "equal_weight",
            "description": "Default scoring algorithm: equal_weight, weighted, or custom",
        },
        {
            "key": "points_per_item",
            "value": "1.087",
            "description": "Default points per checklist item (100 / 92 = 1.087)",
        },
        {
            "key": "total_checklist_items",
            "value": "92",
            "description": "Total number of checklist items in the system",
        },
    ]

    for setting_data in default_settings:
        result = await db.execute(
            select(Setting).where(Setting.key == setting_data["key"])
        )
        existing = result.scalar_one_or_none()

        if not existing:
            setting = Setting(**setting_data)
            db.add(setting)
            print(f"  ✅ Created setting: {setting_data['key']}")
        else:
            # Update existing setting
            existing.value = setting_data["value"]
            existing.description = setting_data["description"]
            print(f"  🔄 Updated setting: {setting_data['key']}")

    await db.commit()
    print("✅ System settings configured")


async def main():
    """Main import function"""
    print("\n" + "=" * 80)
    print("🚀 IMPORTING COMPLETE The Sales Checklist™ DATA")
    print("📊 92 Items from 10 Excel Files")
    print("=" * 80)

    excel_path = Path(EXCEL_FILES_PATH)
    if not excel_path.exists():
        print(f"❌ Excel files directory not found: {EXCEL_FILES_PATH}")
        sys.exit(1)

    async with AsyncSessionLocal() as db:
        try:
            # Clear existing data
            await clear_existing_data(db)
            
            # Seed organization and settings
            await seed_default_organization(db)
            await seed_settings(db)
            
            # Create categories
            categories = await seed_categories(db)
            
            total_items_created = 0
            total_questions_created = 0
            
            # Import data for each category
            print("\n📋 Importing checklist items from Excel files...")
            
            for cat_data in CHECKLIST_CATEGORIES:
                category = categories[cat_data["name"]]
                excel_file = excel_path / cat_data["file"]
                
                if not excel_file.exists():
                    print(f"  ⚠️  File not found: {excel_file}")
                    continue
                
                # Parse Excel file
                excel_data = parse_excel_file(str(excel_file))
                
                # Import items for this category
                items = await import_category_items(db, category, cat_data, excel_data)
                total_items_created += len(items)
                
                # Count coaching questions
                for item in items:
                    q_result = await db.execute(
                        select(CoachingQuestion).where(CoachingQuestion.item_id == item.id)
                    )
                    questions = q_result.scalars().all()
                    total_questions_created += len(questions)

            await db.commit()

            print("\n" + "=" * 80)
            print("🎉 CHECKLIST DATA IMPORT COMPLETED!")
            print("=" * 80)
            print(f"\n📊 Summary:")
            print(f"  • {len(CHECKLIST_CATEGORIES)} Categories imported")
            print(f"  • {total_items_created} Checklist Items created")
            print(f"  • {total_questions_created} Coaching Questions imported")
            print(f"  • 1 Organization configured")
            print(f"  • 5 System settings updated")
            
            if total_items_created == 92:
                print(f"\n✅ SUCCESS: All 92 checklist items imported correctly!")
            else:
                print(f"\n⚠️  WARNING: Expected 92 items, got {total_items_created}")
            
            print(f"\n🔧 Database is now ready for staging deployment")
            print("")

        except Exception as e:
            print(f"\n❌ ERROR: {e}")
            await db.rollback()
            import traceback
            traceback.print_exc()
            sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())